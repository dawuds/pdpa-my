#!/usr/bin/env python3
"""
PDPA Data Protection Audit Work Program (AWP) Generator
Generates: PDPA-DPA-Core-WorkProgram.xlsx

Design principles:
- Act 709 + Amendment A1727 driven
- Two-scope model: organisation-level and processing-activity-level controls
- Compliance assessment (C/PC/NC/NA)
- Legal overlay: 7 Principles, A1727 rights, breach notification, director liability (s133A)
- ROPA (Record of Processing Activities) as the spine for activity-level testing
- Penalty exposure matrix with A1727 enhanced penalties

Reads control library from controls/library.json dynamically.
Backward-compatible with v2.0 schema (generates AWP requirements from keyActivities
when awpRequirements field is absent).
"""

import json
import os
from datetime import date
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=12, bold=True)
BODY_FONT = Font(name="Calibri", size=10)
SMALL_FONT = Font(name="Calibri", size=9, italic=True)
BOLD_BODY = Font(name="Calibri", size=10, bold=True)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LEGAL_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
LEGAL_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CONTEXT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CONTEXT_FONT = Font(name="Calibri", size=9, italic=True)
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_GREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ORG_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ACTIVITY_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
PRINCIPLE_FILL = PatternFill(start_color="E8E0F0", end_color="E8E0F0", fill_type="solid")

# Compliance status fills for conditional formatting
COMPLIANT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NC_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Risk tier fills for context rows
HIGH_RISK_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
MEDIUM_RISK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LOW_RISK_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_json(path):
    with open(path, "r") as f:
        return json.load(f)


def style_header_row(ws, row, max_col, fill=None, font=None):
    f = fill or HEADER_FILL
    fn = font or HEADER_FONT
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = fn
        cell.fill = f
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def write_row(ws, row, values, font=None, fill=None, alignment=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or BODY_FONT
        if fill:
            cell.fill = fill
        cell.alignment = alignment or WRAP
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_compliance_validation(ws, col_letter, start_row, end_row):
    """Add C/PC/NC/NA dropdown validation to compliance column."""
    dv = DataValidation(
        type="list",
        formula1='"C,PC,NC,NA"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid",
        error="Use C, PC, NC, or NA",
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def add_compliance_conditional_formatting(ws, col_letter, start_row, end_row):
    """Add colour coding for C/PC/NC compliance status cells."""
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"C"'], fill=COMPLIANT_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"PC"'], fill=PARTIAL_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"NC"'], fill=NC_FILL)
    )


def freeze_panes(ws, cell_ref="A3"):
    ws.freeze_panes = cell_ref


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
library_data = load_json(os.path.join(REPO_ROOT, "controls", "library.json"))
domains_data = load_json(os.path.join(REPO_ROOT, "controls", "domains.json"))

CONTROLS_LIST = library_data.get("controls", [])

# Build ordered domain list from library.json meta (preserves intended order)
DOMAIN_ORDER = [d["slug"] for d in library_data.get("domains", [])]

# Build domain info lookup from domains.json
DOMAIN_INFO = {}
for slug, info in domains_data.items():
    DOMAIN_INFO[slug] = info

# Group controls by domain
CONTROLS_BY_DOMAIN = OrderedDict()
for slug in DOMAIN_ORDER:
    CONTROLS_BY_DOMAIN[slug] = []
for ctrl in CONTROLS_LIST:
    d = ctrl["domain"]
    if d in CONTROLS_BY_DOMAIN:
        CONTROLS_BY_DOMAIN[d].append(ctrl)
    else:
        CONTROLS_BY_DOMAIN[d] = [ctrl]

# ---------------------------------------------------------------------------
# Placeholder ROPA entries for pre-population
# ---------------------------------------------------------------------------
ROPA_ACTIVITIES = [
    {"id": "PA-001", "purpose": "[Customer Onboarding]", "legal_basis": "[Consent / Contract]",
     "data_categories": "[Name, IC, Contact]", "data_subjects": "[Customers]",
     "retention": "[7 years]", "cross_border": "[Y/N]", "risk_level": "[High/Medium/Low]"},
    {"id": "PA-002", "purpose": "[Employee HR Management]", "legal_basis": "[Contract / Legal Obligation]",
     "data_categories": "[Name, IC, Salary, Health]", "data_subjects": "[Employees]",
     "retention": "[7 years post-employment]", "cross_border": "[Y/N]", "risk_level": "[High/Medium/Low]"},
    {"id": "PA-003", "purpose": "[Marketing Communications]", "legal_basis": "[Consent]",
     "data_categories": "[Name, Email, Preferences]", "data_subjects": "[Prospects/Customers]",
     "retention": "[Until withdrawal]", "cross_border": "[Y/N]", "risk_level": "[High/Medium/Low]"},
    {"id": "PA-004", "purpose": "[CCTV Surveillance]", "legal_basis": "[Legitimate Interest]",
     "data_categories": "[Image, Location]", "data_subjects": "[Visitors/Employees]",
     "retention": "[30 days]", "cross_border": "[N]", "risk_level": "[Medium]"},
    {"id": "PA-005", "purpose": "[Vendor Management]", "legal_basis": "[Contract]",
     "data_categories": "[Name, Contact, Company]", "data_subjects": "[Vendor Representatives]",
     "retention": "[Contract + 7 years]", "cross_border": "[Y/N]", "risk_level": "[Low/Medium]"},
]

# ---------------------------------------------------------------------------
# 7 PDPA Principles
# ---------------------------------------------------------------------------
PDPA_PRINCIPLES = [
    ("General Principle (s6)", "general",
     "Personal data shall not be processed unless consent given or falls within exceptions.",
     ["consent-management"]),
    ("Notice and Choice Principle (s7)", "notice-and-choice",
     "Data subject informed by written notice of purpose, right to request access/correction, class of third parties, choices for limiting processing.",
     ["notice-transparency", "consent-management"]),
    ("Disclosure Principle (s8)", "disclosure",
     "Personal data shall not be disclosed for any purpose other than the purpose disclosed at time of collection or directly related purpose.",
     ["third-party-management", "cross-border-transfer"]),
    ("Security Principle (s9)", "security",
     "Practical steps to protect personal data from loss, misuse, modification, unauthorised access/disclosure/alteration/destruction.",
     ["data-security", "third-party-management"]),
    ("Retention Principle (s10)", "retention",
     "Personal data shall not be kept longer than necessary for fulfilment of the purpose.",
     ["data-lifecycle"]),
    ("Data Integrity Principle (s11)", "data-integrity",
     "Reasonable steps to ensure personal data is accurate, complete, not misleading, and kept up to date.",
     ["data-lifecycle", "data-subject-rights"]),
    ("Access Principle (s12)", "access",
     "Data subject shall be given access to personal data and able to correct inaccurate, incomplete, misleading, or out-of-date data.",
     ["data-subject-rights"]),
]

# ---------------------------------------------------------------------------
# A1727 Rights
# ---------------------------------------------------------------------------
A1727_RIGHTS = [
    ("Right to Withdraw Consent", "s38 / A1727", "consent-management",
     "Data subject may withdraw consent at any time. Withdrawal must be as easy as giving consent.",
     ["Withdrawal mechanism exists on every consent channel",
      "Withdrawal processed within 14 days",
      "Downstream systems and processors notified of withdrawal",
      "Data subject receives confirmation of withdrawal completion"]),
    ("Right to Erasure", "s43 / A1727", "data-lifecycle",
     "Data subject may request deletion of personal data where processing no longer necessary or consent withdrawn.",
     ["Erasure request procedure documented and accessible",
      "Systematic search across all systems holding data subject's data",
      "Erasure confirmation within statutory timeframe",
      "Evidence of erasure retained (without personal data)"]),
    ("Right to Data Portability", "s43A / A1727", "data-subject-rights",
     "Data subject may request personal data in structured, commonly used, machine-readable format.",
     ["Portability request handling procedure exists",
      "Standard export format defined (CSV/JSON/XML)",
      "Response within 21-day statutory deadline",
      "Direct transfer to another data user where technically feasible"]),
    ("Right to Restrict Processing", "s44 / A1727", "consent-management",
     "Data subject may request restriction of processing in specified circumstances.",
     ["Restriction mechanism documented and implemented in systems",
      "Restricted data flagged — processing halted except for storage",
      "Restriction lifted only with data subject agreement or legal basis",
      "Third parties notified of restriction"]),
    ("Right to Object to Automated Decision-Making", "A1727", "governance-accountability",
     "Data subject may object to decisions made solely by automated processing that significantly affect them.",
     ["Inventory of automated decision-making processes maintained",
      "Human review mechanism available for contested decisions",
      "Data subject informed when automated decisions are made",
      "Impact assessment conducted for ADM processes"]),
]

# ---------------------------------------------------------------------------
# Breach notification requirements (s12B)
# ---------------------------------------------------------------------------
BREACH_NOTIFICATION_ITEMS = [
    ("s12B(1)", "Notification to Commissioner", "72 hours",
     "Notify Commissioner within 72 hours of becoming aware of personal data breach.",
     ["Breach response plan with 72-hour notification procedure",
      "Designated breach notification officer identified",
      "Commissioner notification template prepared",
      "Breach assessment criteria for notification trigger defined"]),
    ("s12B(2)", "Notification to Data Subject", "Without undue delay",
     "Notify affected data subjects where breach likely to result in significant harm.",
     ["Harm assessment criteria documented",
      "Data subject notification template prepared (BM and EN)",
      "Communication channels identified for rapid notification",
      "Call centre briefing pack for breach notification calls"]),
    ("Breach Response Plan", "Internal", "Ongoing",
     "Documented breach response plan covering detection, assessment, containment, investigation, notification, and remediation.",
     ["Breach response plan approved by management",
      "RACI matrix for breach response roles",
      "Escalation procedures with timeframes",
      "Post-breach review and lessons learned procedure"]),
    ("Tabletop Exercise", "Internal", "Annual",
     "Regular testing of breach response capabilities through simulation exercises.",
     ["Annual tabletop exercise conducted",
      "72-hour notification walkthrough tested",
      "Exercise results documented with improvement actions",
      "Management briefed on exercise outcomes"]),
]

# ---------------------------------------------------------------------------
# Cross-border transfer mechanisms
# ---------------------------------------------------------------------------
CROSS_BORDER_MECHANISMS = [
    ("Adequacy", "s129(3)(a)", "Transfer to country with adequate data protection laws as determined by Minister.",
     ["List of adequate countries identified",
      "Adequacy assessment documented for each destination",
      "Changes in adequacy status monitored"]),
    ("Contractual Safeguards", "s129(3)(b)", "Standard contractual clauses or binding corporate rules.",
     ["Standard contractual clauses template in use",
      "All cross-border transfers covered by contract",
      "Contract includes PDPA-equivalent obligations",
      "Periodic review of contractual compliance"]),
    ("Binding Corporate Rules", "s129(3)(c)", "Group-wide data protection policies for intra-group transfers.",
     ["BCR document approved",
      "BCR covers all PDPA principles",
      "BCR binding on all group entities",
      "Annual BCR compliance review"]),
    ("Consent", "s129(3)(d)", "Explicit informed consent from data subject for the specific transfer.",
     ["Consent specifically covers cross-border transfer",
      "Destination country disclosed in consent",
      "Consent records maintained with transfer details",
      "Withdrawal mechanism available for transfer consent"]),
]

# ---------------------------------------------------------------------------
# Director liability (s133A)
# ---------------------------------------------------------------------------
S133A_ASSESSMENT = [
    ("Board-level data protection governance",
     "Does the board have formal data protection governance? Is there a DPO with board reporting?",
     "Board charter, DPO appointment, reporting schedule, meeting minutes (12 months)"),
    ("Active vs passive oversight",
     "Do board minutes show active engagement with data protection issues or passive receipt of reports?",
     "Board minutes analysis, decision records, action item tracking"),
    ("DPO appointment and authority",
     "Is there a DPO with adequate authority, resources, and independence? Direct reporting line to board?",
     "DPO appointment letter, job description, reporting structure, resource allocation"),
    ("PDPA training for directors",
     "Have directors been trained on their personal liability under s133A?",
     "Training records, signed acknowledgements, training content"),
    ("Response to known risks",
     "When audits or assessments identified gaps, did the board direct and fund remediation?",
     "Audit reports, remediation tracker, budget approvals, board action items"),
    ("Data protection impact assessments",
     "Are DPIAs conducted for high-risk processing and reviewed by management?",
     "DPIA register, DPIA reports, management sign-off records"),
    ("Compliance monitoring programme",
     "Is there an ongoing compliance monitoring programme with board reporting?",
     "Compliance reports, monitoring schedule, board briefing records"),
    ("Due diligence documentation",
     "Could each director demonstrate 'all due diligence' per s133A(2) if required?",
     "Composite assessment of above evidence"),
]

# ---------------------------------------------------------------------------
# Penalty exposure matrix
# ---------------------------------------------------------------------------
PENALTY_MATRIX = [
    ("s5(1)", "Processing without registration", "RM500,000", "3 years", "Both",
     "Compoundable", "A1727 enhanced"),
    ("s16(4)", "Failure to renew registration", "RM250,000", "2 years", "Both",
     "Compoundable", "Original"),
    ("s16A(5)", "Processing while registration suspended/revoked", "RM500,000", "3 years", "Both",
     "Compoundable", "A1727 enhanced"),
    ("s5A(2)", "Non-compliance with General Principle (s6)", "RM500,000", "3 years", "Both",
     "Non-compoundable", "A1727 new"),
    ("s5A(2)", "Non-compliance with Notice and Choice Principle (s7)", "RM500,000", "3 years", "Both",
     "Non-compoundable", "A1727 new"),
    ("s5A(2)", "Non-compliance with Disclosure Principle (s8)", "RM500,000", "3 years", "Both",
     "Non-compoundable", "A1727 new"),
    ("s5A(2)", "Non-compliance with Security Principle (s9)", "RM500,000", "3 years", "Both",
     "Non-compoundable", "A1727 new"),
    ("s5A(2)", "Non-compliance with Retention Principle (s10)", "RM300,000", "2 years", "Both",
     "Non-compoundable", "A1727 new"),
    ("s5A(2)", "Non-compliance with Data Integrity Principle (s11)", "RM300,000", "2 years", "Both",
     "Non-compoundable", "A1727 new"),
    ("s5A(2)", "Non-compliance with Access Principle (s12)", "RM300,000", "2 years", "Both",
     "Non-compoundable", "A1727 new"),
    ("s12B(5)", "Failure to notify breach to Commissioner (72 hours)", "RM250,000", "2 years", "Both",
     "Compoundable", "A1727 new"),
    ("s12B(6)", "Failure to notify data subject of breach", "RM250,000", "2 years", "Both",
     "Compoundable", "A1727 new"),
    ("s30(7)", "Failure to comply with data access request (21 days)", "RM200,000", "1 year", "Both",
     "Compoundable", "Original"),
    ("s34(5)", "Failure to comply with correction request", "RM200,000", "1 year", "Both",
     "Compoundable", "Original"),
    ("s43(3)", "Failure to cease direct marketing", "RM200,000", "1 year", "Both",
     "Compoundable", "Original"),
    ("s43A(3)", "Failure to provide data portability", "RM300,000", "2 years", "Both",
     "Compoundable", "A1727 new"),
    ("s44(3)", "Failure to restrict processing on request", "RM300,000", "2 years", "Both",
     "Compoundable", "A1727 new"),
    ("s129(1)", "Unauthorised cross-border transfer", "RM300,000", "2 years", "Both",
     "Non-compoundable", "Enhanced"),
    ("s133A", "Director/officer liability (consent/connivance/neglect)", "Same as body corporate", "Same", "Personal",
     "Per underlying offence", "A1727 new"),
    ("s130", "Obstruction of Commissioner/officer", "RM200,000", "1 year", "Both",
     "Non-compoundable", "Original"),
    ("s131", "Providing false information", "RM300,000", "2 years", "Both",
     "Non-compoundable", "Original"),
]


# ---------------------------------------------------------------------------
# PBC items
# ---------------------------------------------------------------------------
PBC_ITEMS = [
    # Consent Management
    ("consent-management", "Data protection registration certificate", "Essential"),
    ("consent-management", "Consent form templates (all channels — web, app, paper, call centre)", "Essential"),
    ("consent-management", "Consent management system configuration and screenshots", "Essential"),
    ("consent-management", "Consent records sample (20 records across channels)", "Important"),
    ("consent-management", "Consent withdrawal requests log (12 months)", "Important"),
    ("consent-management", "Consent validity audit report (if available)", "Important"),
    ("consent-management", "Sensitive personal data consent forms (s29 categories)", "Essential"),
    ("consent-management", "Consent re-consent campaign records (if conducted)", "Important"),
    # Data Subject Rights
    ("data-subject-rights", "DSAR intake procedure and forms", "Essential"),
    ("data-subject-rights", "DSAR register / case management system extract (12 months)", "Essential"),
    ("data-subject-rights", "Sample of 10 completed DSARs with evidence of response", "Essential"),
    ("data-subject-rights", "DSAR response time analysis report", "Important"),
    ("data-subject-rights", "Data correction request samples with propagation evidence", "Important"),
    ("data-subject-rights", "Direct marketing suppression list", "Important"),
    ("data-subject-rights", "Data portability request handling procedure (A1727)", "Essential"),
    # Notice and Transparency
    ("notice-transparency", "Privacy notice — current version (BM and EN)", "Essential"),
    ("notice-transparency", "Privacy notice version history (last 3 versions)", "Essential"),
    ("notice-transparency", "Privacy notice delivery records (web analytics, acknowledgement logs)", "Important"),
    ("notice-transparency", "CCTV and monitoring disclosure notices", "Important"),
    ("notice-transparency", "Privacy notice review schedule and last review minutes", "Important"),
    ("notice-transparency", "s7 notice checklist / compliance mapping", "Essential"),
    # Data Security
    ("data-security", "Information security policy (current approved version)", "Essential"),
    ("data-security", "Data classification scheme and data inventory", "Essential"),
    ("data-security", "Encryption standards and deployment evidence", "Essential"),
    ("data-security", "Access control policy and current user access matrix", "Essential"),
    ("data-security", "Vulnerability assessment reports (latest 2)", "Important"),
    ("data-security", "Penetration test reports (latest)", "Important"),
    ("data-security", "Security awareness training records (12 months)", "Important"),
    ("data-security", "Security incident log (12 months)", "Important"),
    ("data-security", "Physical security procedures and access logs", "Important"),
    # Data Lifecycle
    ("data-lifecycle", "Data retention policy and schedule", "Essential"),
    ("data-lifecycle", "Record of Processing Activities (ROPA)", "Essential"),
    ("data-lifecycle", "Data disposal/destruction records (12 months)", "Essential"),
    ("data-lifecycle", "Data quality review reports", "Important"),
    ("data-lifecycle", "Storage system inventory with retention enforcement config", "Important"),
    ("data-lifecycle", "DPIA reports for high-risk processing (if available)", "Important"),
    # Breach Response
    ("breach-response", "Data breach response plan", "Essential"),
    ("breach-response", "Breach notification templates (Commissioner and data subject)", "Essential"),
    ("breach-response", "Breach log / register (12 months)", "Essential"),
    ("breach-response", "72-hour notification tabletop exercise results", "Essential"),
    ("breach-response", "Breach response team roster and contact list", "Important"),
    ("breach-response", "Forensic investigation tools inventory", "Important"),
    ("breach-response", "Post-breach review reports (if any breaches occurred)", "Important"),
    # Cross-Border Transfer
    ("cross-border-transfer", "Cross-border transfer register / inventory", "Essential"),
    ("cross-border-transfer", "Transfer impact assessments for each destination", "Essential"),
    ("cross-border-transfer", "Standard contractual clauses templates in use", "Essential"),
    ("cross-border-transfer", "Binding corporate rules (if applicable)", "Important"),
    ("cross-border-transfer", "Cross-border consent records (if consent is transfer basis)", "Important"),
    ("cross-border-transfer", "Adequacy assessment documentation", "Important"),
    # Third-Party Management
    ("third-party-management", "Data processing agreements (sample of 10)", "Essential"),
    ("third-party-management", "Processor due diligence assessment records", "Essential"),
    ("third-party-management", "Processor register / inventory", "Essential"),
    ("third-party-management", "Processor audit reports or certifications", "Important"),
    ("third-party-management", "Sub-processor approval records", "Important"),
    ("third-party-management", "Processor breach notification records", "Important"),
    # Governance and Accountability
    ("governance-accountability", "DPO appointment letter and job description", "Essential"),
    ("governance-accountability", "Data protection policy framework", "Essential"),
    ("governance-accountability", "Board/management data protection briefing records (12 months)", "Essential"),
    ("governance-accountability", "Staff training plan and completion records", "Essential"),
    ("governance-accountability", "Privacy by design procedure / DPIA procedure", "Essential"),
    ("governance-accountability", "Compliance monitoring reports (quarterly)", "Important"),
    ("governance-accountability", "Data protection audit reports (internal/external)", "Important"),
    ("governance-accountability", "Commissioner correspondence file", "Important"),
    ("governance-accountability", "Automated decision-making inventory (A1727)", "Important"),
    # Registration and Compliance
    ("registration-compliance", "Data user registration certificate", "Essential"),
    ("registration-compliance", "Registration renewal records", "Essential"),
    ("registration-compliance", "Commissioner correspondence and inspection records", "Essential"),
    ("registration-compliance", "Code of Practice compliance assessment (sector-specific)", "Important"),
    ("registration-compliance", "Regulatory change monitoring log", "Important"),
    ("registration-compliance", "Annual fee payment records", "Important"),
]


# ---------------------------------------------------------------------------
# Build AWP requirements from library.json
# ---------------------------------------------------------------------------
def build_awp_requirements(controls_list):
    """
    Build AWP requirements from control library.
    If a control has awpRequirements (v3.0), use them directly.
    Otherwise, generate 2-3 requirements per control from keyActivities.
    """
    requirements_by_domain = OrderedDict()
    req_counter = {}

    for ctrl in controls_list:
        domain = ctrl["domain"]
        if domain not in requirements_by_domain:
            requirements_by_domain[domain] = []
            req_counter[domain] = 0

        awp_reqs = ctrl.get("awpRequirements", [])
        if awp_reqs:
            # v3.0 path — use structured AWP requirements
            for req in awp_reqs:
                req_counter[domain] += 1
                requirements_by_domain[domain].append({
                    "req_id": req.get("id", f"AWP-{domain[:4].upper()}-{req_counter[domain]:03d}"),
                    "description": req.get("description", ""),
                    "scope": req.get("scope", "org"),
                    "sections": ctrl.get("sections", []),
                    "control_slug": ctrl["slug"],
                    "control_name": ctrl["name"],
                })
        else:
            # v2.0 path — generate from keyActivities
            activities = ctrl.get("keyActivities", [])
            # Take up to 3 activities as requirements
            selected = activities[:3] if len(activities) >= 3 else activities
            if not selected:
                # Fallback: use control description as a single requirement
                selected = [ctrl.get("description", ctrl["name"])]

            for activity in selected:
                req_counter[domain] += 1
                # Determine scope based on control type and layer
                scope = "org"
                layer = ctrl.get("layer", "process")
                ctrl_type = ctrl.get("type", "")
                if layer == "technology" or ctrl_type in ("technical", "detective"):
                    scope = "activity"

                requirements_by_domain[domain].append({
                    "req_id": f"AWP-{domain[:4].upper()}-{req_counter[domain]:03d}",
                    "description": activity,
                    "scope": scope,
                    "sections": ctrl.get("sections", []),
                    "control_slug": ctrl["slug"],
                    "control_name": ctrl["name"],
                })

    return requirements_by_domain


AWP_REQUIREMENTS = build_awp_requirements(CONTROLS_LIST)


# ---------------------------------------------------------------------------
# Sheet creators
# ---------------------------------------------------------------------------
def create_cover_sheet(wb):
    """Sheet 1: Cover & Scope."""
    ws = wb.active
    ws.title = "1. Cover & Scope"
    set_col_widths(ws, [30, 55, 35])

    ws.merge_cells("A1:C1")
    ws["A1"] = "PDPA Data Protection Audit — Core Work Program"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:C2")
    ws["A2"] = "Personal Data Protection Act 2010 (Act 709) + Amendment A1727 — Compliance Assessment"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)

    row = 4
    fields = [
        ("Entity Name", ""),
        ("Registration Number", "[Data User Registration No.]"),
        ("Sector", "[e.g., Banking, Healthcare, Telecommunications, Insurance]"),
        ("Applicable Code of Practice", "[e.g., General CoP 2022, Sector-specific CoP]"),
        ("Audit Period", ""),
        ("Assessment Type", "Compliance (Mandatory) + Risk-Based (Mandatory)"),
        ("Approaches Elected", "[Compliance-Based / Risk-Based / Control-Based / Technical Testing]"),
        ("Engagement Partner", ""),
        ("Lead Auditor", ""),
        ("Legal Adviser", "[For legal interpretation — s133A, cross-border, breach notification]"),
        ("DPO Contact", "[Entity DPO name and contact]"),
        ("Processing Activities Count", "[From ROPA — Sheet 2]"),
        ("Report Due Date", ""),
    ]
    headers = ["Field", "Value", "Notes"]
    write_row(ws, row, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    for label, default in fields:
        write_row(ws, row, [label, default, ""])
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    # Sampling
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Sampling Rationale"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    sampling = [
        ("Processing Activities in Scope", "[From Sheet 2]", ""),
        ("Activity-Level Sampling", "1-5: Full population | 6-15: All high-risk + sample | 16+: Risk-based sample", ""),
        ("Sampling Method", "Judgmental — appropriate for compliance engagements", ""),
        ("Sample Size Justification", "", ""),
    ]
    for label, default, notes in sampling:
        write_row(ws, row, [label, default, notes])
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    # Population-based sampling table
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Sampling Methodology"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    write_row(ws, row, ["Population Size", "Standard Risk Sample", "High Risk Sample"],
              font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    pop_samples = [
        ("1-5", "All (full coverage)", "All (full coverage)"),
        ("6-15", "3", "5"),
        ("16-50", "5", "8"),
        ("51-100", "8", "12"),
        ("100+", "10", "15 (with stratification)"),
    ]
    for pop, std, high in pop_samples:
        write_row(ws, row, [pop, std, high])
        row += 1

    # Time-based coverage
    row += 1
    write_row(ws, row, ["Evidence Type", "Coverage Period", "Rationale"],
              font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    time_coverage = [
        ("Board/management reports", "Last 4 quarters", "Full annual governance cycle"),
        ("Operational records (DSARs, breaches, consent)", "Last 6-12 months", "Recent operating effectiveness"),
        ("Annual processes (DPIA, training, audit)", "Last 12 months", "Full cycle"),
        ("Policies and notices", "Current version + prior", "Change tracking and currency"),
        ("System configurations (access, encryption)", "Point in time", "Current state verification"),
    ]
    for etype, period, rationale in time_coverage:
        write_row(ws, row, [etype, period, rationale])
        row += 1

    # Compliance rating key
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Compliance Rating Key"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    ratings = [
        ("C — Compliant", "Requirement fully met: policy/control exists, approved, applied, current, evidenced"),
        ("PC — Partially Compliant", "Some elements exist but gaps remain; started but not completed; covers some but not all processing activities"),
        ("NC — Non-Compliant", "Not met or no evidence; entity not aware of requirement; fundamentally inadequate"),
        ("NA — Not Applicable", "Justified exclusion (e.g., no cross-border transfers) — must provide written justification"),
    ]
    for code, desc in ratings:
        write_row(ws, row, [code, desc, ""])
        row += 1

    # Disclaimer
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "Disclaimer"
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = (
        "This work program template is constructed-indicative for educational purposes. "
        "It must be tailored to each entity's specific context, sector, and processing activities. "
        "Not a verbatim extract of any official PDPC document. "
        "AI-generated content — requires professional review before audit/compliance use."
    )
    ws[f"A{row}"].font = SMALL_FONT
    ws[f"A{row}"].alignment = WRAP

    freeze_panes(ws, "A5")


def create_ropa_register(wb):
    """Sheet 2: Data Processing Activity Register (ROPA)."""
    ws = wb.create_sheet("2. ROPA Register")
    set_col_widths(ws, [12, 30, 20, 25, 20, 18, 14, 14, 14, 30])

    ws.merge_cells("A1:J1")
    ws["A1"] = "Record of Processing Activities (ROPA) — Audit Scope Definition"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        "This register defines the processing activities in scope. "
        "Activity-level controls in domain worksheets reference activities listed here. "
        "Populate from entity's ROPA or conduct mapping exercise during fieldwork."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP

    row = 4
    headers = [
        "Activity ID", "Processing Purpose", "Legal Basis",
        "Data Categories", "Data Subjects", "Retention Period",
        "Cross-Border?", "Risk Level", "In Audit Scope?", "Notes"
    ]
    write_row(ws, row, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1

    for act in ROPA_ACTIVITIES:
        write_row(ws, row, [
            act["id"], act["purpose"], act["legal_basis"],
            act["data_categories"], act["data_subjects"], act["retention"],
            act["cross_border"], act["risk_level"], "Y/N", ""
        ])
        row += 1

    # Extra blank rows
    for i in range(15):
        write_row(ws, row, [f"PA-{len(ROPA_ACTIVITIES) + i + 1:03d}"] + [""] * 9)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = (
        "Note: Processing activities NOT in the ROPA but discovered during audit should be "
        "flagged as a finding (governance-accountability domain). All activity-level rows "
        "in domain worksheets reference activities listed here."
    )
    ws[f"A{row}"].font = SMALL_FONT
    ws[f"A{row}"].alignment = WRAP

    # Add risk level validation
    dv = DataValidation(
        type="list", formula1='"High,Medium,Low"', allow_blank=True
    )
    dv.sqref = f"H5:H{row - 2}"
    ws.add_data_validation(dv)

    freeze_panes(ws, "A5")


def create_domain_worksheet(wb, domain_slug, domain_info, controls, sheet_num):
    """Sheets 3-12: One per domain."""
    domain_name = domain_info.get("name", domain_slug.replace("-", " ").title())
    sheet_title = f"{sheet_num}. {domain_name}"
    if len(sheet_title) > 31:
        # Truncate to fit Excel's 31-char limit
        sheet_title = sheet_title[:31]

    ws = wb.create_sheet(sheet_title)

    max_col = 8
    set_col_widths(ws, [14, 45, 12, 14, 14, 18, 14, 30])

    # Title
    ws.merge_cells(f"A1:{get_column_letter(max_col)}1")
    ws["A1"] = f"{domain_name} ({len(controls)} controls)"
    ws["A1"].font = TITLE_FONT

    # Domain metadata
    related_sections = domain_info.get("relatedSections", [])
    related_principles = domain_info.get("relatedPrinciples", [])
    ws.merge_cells(f"A2:{get_column_letter(max_col)}2")
    ws["A2"] = (
        f"PDPA Sections: {', '.join(related_sections) if related_sections else 'General'} | "
        f"Principles: {', '.join(related_principles) if related_principles else 'General'} | "
        f"Controls: {len(controls)}"
    )
    ws["A2"].font = SMALL_FONT

    row = 4
    reqs = AWP_REQUIREMENTS.get(domain_slug, [])

    # Group requirements by control
    control_groups = OrderedDict()
    for req in reqs:
        slug = req["control_slug"]
        if slug not in control_groups:
            # Find the full control object
            ctrl_obj = None
            for c in controls:
                if c["slug"] == slug:
                    ctrl_obj = c
                    break
            control_groups[slug] = {"control": ctrl_obj, "reqs": []}
        control_groups[slug]["reqs"].append(req)

    # --- Write header ---
    headers = [
        "Req ID", "Requirement Description", "Scope\n(Org/Act)",
        "PDPA Section", "Compliance\n(C/PC/NC/NA)", "Evidence Ref",
        "Auditor Notes", "Finding Ref"
    ]
    write_row(ws, row, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    data_start = row

    for ctrl_slug, group in control_groups.items():
        ctrl = group["control"]

        # --- Context row: control-level guidance ---
        if ctrl:
            # Check for v3.0 fields first, fallback to description-based context
            obj = ctrl.get("controlObjective", "")
            wgll = ctrl.get("whatGoodLooksLike", "")
            risk = ctrl.get("keyRiskIfAbsent", "")
            sampling = ctrl.get("samplingGuidance", "")
            tier = ctrl.get("riskTier", "")

            if obj or wgll or risk:
                # v3.0 context
                context_text = f"WHY: {obj}\nWHAT GOOD LOOKS LIKE: {wgll}\nKEY RISK IF ABSENT: {risk}"
                if sampling:
                    context_text += f"\nSAMPLING: {sampling}"
            else:
                # v2.0 fallback: use description + maturity levels
                maturity = ctrl.get("maturity", {})
                mature_desc = maturity.get("mature", "")
                context_text = (
                    f"OBJECTIVE: {ctrl.get('description', '')}\n"
                    f"WHAT GOOD LOOKS LIKE: {mature_desc}"
                )

            # Determine fill based on risk tier
            if tier == "critical" or tier == "high":
                tier_fill = HIGH_RISK_FILL
            elif tier == "low" or tier == "conditional":
                tier_fill = LOW_RISK_FILL
            else:
                tier_fill = CONTEXT_FILL

            # Control name header
            tier_label = f"  [{tier.upper()}]" if tier else ""
            ws.merge_cells(f"A{row}:{get_column_letter(max_col)}{row}")
            ws[f"A{row}"] = f"{ctrl['name']}{tier_label}"
            ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{row}"].fill = tier_fill
            ws[f"A{row}"].alignment = WRAP
            ws[f"A{row}"].border = THIN_BORDER
            row += 1

            # Context detail row
            ws.merge_cells(f"A{row}:{get_column_letter(max_col)}{row}")
            ws[f"A{row}"] = context_text
            ws[f"A{row}"].font = CONTEXT_FONT
            ws[f"A{row}"].fill = CONTEXT_FILL
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws[f"A{row}"].border = THIN_BORDER
            ws.row_dimensions[row].height = 55
            row += 1

        # --- Requirement rows ---
        for req in group["reqs"]:
            sections_str = ", ".join(req.get("sections", []))
            scope_label = "Org" if req["scope"] == "org" else "Act"
            write_row(ws, row, [
                req["req_id"], req["description"], scope_label,
                sections_str, "", "", "", ""
            ])
            row += 1

    data_end = row - 1

    # Apply compliance validation and conditional formatting
    if data_start <= data_end:
        add_compliance_validation(ws, "E", data_start, data_end)
        add_compliance_conditional_formatting(ws, "E", data_start, data_end)

    # --- Domain summary section ---
    row += 2
    ws.merge_cells(f"A{row}:{get_column_letter(max_col)}{row}")
    ws[f"A{row}"] = f"Domain Summary — {domain_name}"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1

    summary_labels = [
        "Total requirements assessed",
        "Compliant (C) count",
        "Partially Compliant (PC) count",
        "Non-Compliant (NC) count",
        "Not Applicable (NA) count",
        "Domain compliance profile",
        "Key findings (reference)",
        "Auditor domain narrative",
    ]
    for label in summary_labels:
        write_row(ws, row, [label, "", ""] + [""] * (max_col - 3))
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    freeze_panes(ws, "A5")


def create_principles_sheet(wb):
    """Sheet 13: 7 Principles Assessment."""
    ws = wb.create_sheet("13. 7 Principles")
    set_col_widths(ws, [35, 40, 30, 18, 30, 30])

    ws.merge_cells("A1:F1")
    ws["A1"] = "PDPA 7 Principles — Compliance Assessment"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Cross-cutting assessment of compliance with each of the 7 data protection principles. "
        "This summarises findings from domain worksheets through the lens of each principle."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP

    row = 4
    headers = [
        "Principle", "Description", "Control Coverage\n(domains)",
        "Compliance\n(C/PC/NC)", "Evidence Summary", "Gaps / Observations"
    ]
    write_row(ws, row, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    data_start = row

    for principle_name, _slug, description, related_domains in PDPA_PRINCIPLES:
        domain_names = ", ".join(
            DOMAIN_INFO.get(d, {}).get("name", d) for d in related_domains
        )
        write_row(ws, row, [
            principle_name, description, domain_names, "", "", ""
        ])
        ws.row_dimensions[row].height = 50
        row += 1

    data_end = row - 1
    add_compliance_validation(ws, "D", data_start, data_end)
    add_compliance_conditional_formatting(ws, "D", data_start, data_end)

    # Summary
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Principles Compliance Summary"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    for label in ["Principles rated Compliant", "Principles rated Partially Compliant",
                   "Principles rated Non-Compliant", "Overall principles assessment narrative"]:
        write_row(ws, row, [label, "", "", "", "", ""])
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    freeze_panes(ws, "A5")


def create_a1727_rights_sheet(wb):
    """Sheet 14: A1727 Rights Compliance."""
    ws = wb.create_sheet("14. A1727 Rights")
    set_col_widths(ws, [30, 14, 35, 18, 18, 18, 30])

    ws.merge_cells("A1:G1")
    ws["A1"] = "Amendment A1727 — Enhanced Rights Compliance Assessment"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "Assessment of compliance with new and enhanced data subject rights "
        "introduced by Amendment Act A1727. Each right is tested for mechanism existence, "
        "operational effectiveness, and evidence of fulfilment."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP

    row = 4
    for right_name, section, domain, description, tests in A1727_RIGHTS:
        # Right header
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = f"{right_name} ({section})"
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].fill = PRINCIPLE_FILL
        ws[f"A{row}"].border = THIN_BORDER
        row += 1

        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = description
        ws[f"A{row}"].font = CONTEXT_FONT
        ws[f"A{row}"].fill = CONTEXT_FILL
        ws[f"A{row}"].alignment = WRAP
        ws[f"A{row}"].border = THIN_BORDER
        row += 1

        # Test headers
        test_headers = [
            "Test Item", "PDPA Section", "Test Description",
            "Mechanism\nExists?", "Tested?",
            "Response Time\nCompliant?", "Evidence / Notes"
        ]
        write_row(ws, row, test_headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
        row += 1
        test_start = row

        for i, test_desc in enumerate(tests, 1):
            write_row(ws, row, [
                f"{right_name[:3]}-T{i:02d}", section, test_desc,
                "", "", "", ""
            ])
            row += 1

        # Add Y/N validation for mechanism and tested columns
        for col_letter in ["D", "E", "F"]:
            dv = DataValidation(
                type="list", formula1='"Y,N,NA"', allow_blank=True
            )
            dv.sqref = f"{col_letter}{test_start}:{col_letter}{row - 1}"
            ws.add_data_validation(dv)

        row += 1

    # Summary
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "A1727 Rights Compliance Summary"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    for label in ["Rights with mechanisms in place", "Rights tested and operational",
                   "Rights with response time compliance", "Overall A1727 readiness assessment"]:
        write_row(ws, row, [label, "", "", "", "", "", ""])
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    freeze_panes(ws, "A4")


def create_breach_notification_sheet(wb):
    """Sheet 15: Breach Notification Readiness."""
    ws = wb.create_sheet("15. Breach Notification")
    set_col_widths(ws, [30, 14, 14, 40, 18, 30])

    ws.merge_cells("A1:F1")
    ws["A1"] = "Breach Notification Readiness Assessment (s12B)"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Amendment A1727 introduced mandatory 72-hour breach notification to the Commissioner (s12B). "
        "This sheet assesses the entity's readiness to detect, assess, and notify within the statutory timeframe."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP

    row = 4
    for req_ref, req_name, timeframe, description, tests in BREACH_NOTIFICATION_ITEMS:
        # Requirement header
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"{req_name} ({req_ref}) — {timeframe}"
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].fill = ACTIVITY_FILL
        ws[f"A{row}"].border = THIN_BORDER
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = description
        ws[f"A{row}"].font = CONTEXT_FONT
        ws[f"A{row}"].fill = CONTEXT_FILL
        ws[f"A{row}"].alignment = WRAP
        ws[f"A{row}"].border = THIN_BORDER
        row += 1

        # Test items
        test_headers = ["Test Item", "Section", "Timeframe", "Assessment", "Status\n(C/PC/NC/NA)", "Evidence / Notes"]
        write_row(ws, row, test_headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
        row += 1
        data_start = row

        for i, test_desc in enumerate(tests, 1):
            write_row(ws, row, [
                f"BR-T{i:02d}", req_ref, timeframe, test_desc, "", ""
            ])
            row += 1

        add_compliance_validation(ws, "E", data_start, row - 1)
        add_compliance_conditional_formatting(ws, "E", data_start, row - 1)
        row += 1

    # Summary
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Breach Notification Readiness Summary"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    for label in [
        "Breach response plan exists and approved?",
        "72-hour notification procedure documented?",
        "Notification templates prepared (Commissioner and data subject)?",
        "Tabletop exercise conducted within last 12 months?",
        "Overall breach notification readiness",
    ]:
        write_row(ws, row, [label, "", "", "", "", ""])
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    freeze_panes(ws, "A4")


def create_cross_border_sheet(wb):
    """Sheet 16: Cross-Border Transfer Assessment."""
    ws = wb.create_sheet("16. Cross-Border Transfer")
    set_col_widths(ws, [25, 14, 40, 18, 14, 30])

    ws.merge_cells("A1:F1")
    ws["A1"] = "Cross-Border Data Transfer Assessment (s129)"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Assessment of each cross-border transfer mechanism used by the entity. "
        "PDPA s129 restricts transfer of personal data outside Malaysia unless conditions met."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP

    row = 4
    for mechanism, section, description, tests in CROSS_BORDER_MECHANISMS:
        # Mechanism header
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"{mechanism} ({section})"
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].fill = PRINCIPLE_FILL
        ws[f"A{row}"].border = THIN_BORDER
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = description
        ws[f"A{row}"].font = CONTEXT_FONT
        ws[f"A{row}"].fill = CONTEXT_FILL
        ws[f"A{row}"].alignment = WRAP
        ws[f"A{row}"].border = THIN_BORDER
        row += 1

        test_headers = [
            "Test Item", "Section", "Assessment Criteria",
            "Status\n(C/PC/NC/NA)", "In Use?", "Evidence / Notes"
        ]
        write_row(ws, row, test_headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
        row += 1
        data_start = row

        for i, test_desc in enumerate(tests, 1):
            write_row(ws, row, [
                f"CB-{mechanism[:3].upper()}-T{i:02d}", section, test_desc,
                "", "", ""
            ])
            row += 1

        add_compliance_validation(ws, "D", data_start, row - 1)
        add_compliance_conditional_formatting(ws, "D", data_start, row - 1)
        row += 1

    # Per-transfer assessment
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Per-Transfer Assessment (one row per active cross-border transfer)"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1

    transfer_headers = [
        "Transfer ID", "Destination\nCountry", "Recipient / Purpose",
        "Mechanism\nUsed", "Status\n(C/PC/NC)", "Notes"
    ]
    write_row(ws, row, transfer_headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    data_start = row

    for i in range(10):
        write_row(ws, row, [f"CBT-{i + 1:03d}", "", "", "", "", ""])
        row += 1

    add_compliance_validation(ws, "E", data_start, row - 1)
    add_compliance_conditional_formatting(ws, "E", data_start, row - 1)

    freeze_panes(ws, "A4")


def create_director_liability_sheet(wb):
    """Sheet 17: Director Liability (s133A)."""
    ws = wb.create_sheet("17. Director Liability")
    set_col_widths(ws, [35, 50, 40, 22, 40])

    ws.merge_cells("A1:E1")
    ws["A1"] = "Section 133A — Director and Officer Personal Liability Assessment"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = LEGAL_FILL

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "Amendment A1727 introduced s133A creating personal liability for directors/officers "
        "via consent, connivance, or NEGLECT. Reverse burden of proof — officers must prove due diligence. "
        "This assessment provides input for legal opinion on director exposure."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP

    row = 4
    # Liability framework
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Liability Triggers (s133A(1))"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    triggers = [
        ("CONSENT", "Director knew about and approved the breach / non-compliance"),
        ("CONNIVANCE", "Director was aware of the breach but did nothing to stop it"),
        ("NEGLECT", "Director failed to exercise proper oversight — NO knowledge/intent required"),
    ]
    for trigger, desc in triggers:
        write_row(ws, row, [trigger, desc, "", "", ""])
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = (
        "Defence Requirement (s133A(2)): Director must prove offence committed WITHOUT "
        "their knowledge/consent/connivance AND that they exercised ALL DUE DILIGENCE."
    )
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True, italic=True)
    ws[f"A{row}"].fill = YELLOW_FILL
    ws[f"A{row}"].alignment = WRAP

    row += 2
    headers = [
        "Assessment Area", "Key Questions", "Expected Evidence",
        "Assessment\n(Strong/Adequate/\nWeak/Insufficient)", "Assessor Notes"
    ]
    write_row(ws, row, headers, font=LEGAL_HEADER_FONT, fill=LEGAL_FILL, alignment=CENTER)
    row += 1

    for area, questions, evidence in S133A_ASSESSMENT:
        write_row(ws, row, [area, questions, evidence, "", ""])
        row += 1

    # Officer-level assessment
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Officer-Level Assessment (one row per director/officer)"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    officer_headers = [
        "Officer Name & Title", "Role in Data Protection Oversight",
        "PDPA Training Completed?",
        "Defence Adequacy\n(Strong/Adequate/Weak/Insufficient)",
        "Notes / Recommendations"
    ]
    write_row(ws, row, officer_headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    for _ in range(8):
        write_row(ws, row, [""] * 5)
        row += 1

    freeze_panes(ws, "A4")


def create_penalty_exposure_sheet(wb):
    """Sheet 18: Penalty Exposure Matrix."""
    ws = wb.create_sheet("18. Penalty Exposure")
    set_col_widths(ws, [14, 40, 18, 14, 14, 18, 16, 18, 30])

    ws.merge_cells("A1:I1")
    ws["A1"] = "Penalty Exposure Matrix — PDPA Act 709 + A1727"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = LEGAL_FILL

    ws.merge_cells("A2:I2")
    ws["A2"] = (
        "Quantifies financial and criminal exposure per offence. "
        "A1727 enhanced penalties flagged. Non-compoundable offences require prosecution."
    )
    ws["A2"].font = SMALL_FONT

    row = 4
    headers = [
        "Section", "Offence", "Max Fine",
        "Imprisonment", "Liability\n(Individual/\nCorporate/Both)", "Compoundable?",
        "A1727 Status", "Current\nExposure\n(H/M/L/None)", "Linked Findings"
    ]
    write_row(ws, row, headers, font=LEGAL_HEADER_FONT, fill=LEGAL_FILL, alignment=CENTER)
    row += 1

    for p in PENALTY_MATRIX:
        section, offence, fine, prison, liability, compound, a1727_status = p
        write_row(ws, row, [
            section, offence, fine, prison, liability,
            compound, a1727_status, "", ""
        ])
        # Highlight non-compoundable
        if compound == "Non-compoundable":
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
                )
        # Highlight A1727 new
        if "A1727" in a1727_status:
            ws.cell(row=row, column=7).font = Font(name="Calibri", size=10, bold=True, color="8B0000")
        row += 1

    # Aggregate summary
    row += 2
    ws.merge_cells(f"A{row}:I{row}")
    ws[f"A{row}"] = "Aggregate Exposure Summary"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    summary_items = [
        ("Total offences with current exposure", ""),
        ("Total maximum fine exposure", ""),
        ("Non-compoundable offences identified", ""),
        ("Maximum imprisonment exposure (individual)", ""),
        ("A1727 enhanced penalty exposure count", ""),
        ("Director/officer personal liability exposure", ""),
    ]
    for label, val in summary_items:
        write_row(ws, row, [label, val] + [""] * 7)
        ws.cell(row=row, column=1).font = BOLD_BODY
        row += 1

    freeze_panes(ws, "A5")


def create_finding_register(wb):
    """Sheet 19: Finding Register."""
    ws = wb.create_sheet("19. Finding Register")
    set_col_widths(ws, [14, 30, 18, 14, 40, 35, 30, 35, 35, 14, 18, 18, 14])

    ws.merge_cells("A1:M1")
    ws["A1"] = "Finding Register — 5 C's Framework"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:M2")
    ws["A2"] = (
        "Each finding follows the 5 C's structure: Condition, Criteria, Cause, Consequence, Corrective Action. "
        "Findings are cross-referenced to domain worksheet rows."
    )
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = WRAP

    row = 4
    headers = [
        "Finding ID", "Title", "Domain", "PDPA Section",
        "Condition\n(What was found)", "Criteria\n(PDPA / CoP reference)",
        "Cause\n(Root cause)", "Consequence\n(Impact + penalty exposure)",
        "Corrective Action\n(Recommendation)",
        "Severity\n(C/H/M/L)", "Status\n(Open/In Progress/\nRemediated/Closed)",
        "Target Date", "Auditor"
    ]
    write_row(ws, row, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1

    # Pre-fill 30 blank rows
    for i in range(1, 31):
        finding_id = f"PDPA-{date.today().year}-{i:03d}"
        write_row(ws, row, [finding_id] + [""] * 12)
        row += 1

    # Severity validation
    dv = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True
    )
    dv.sqref = f"J5:J{row - 1}"
    ws.add_data_validation(dv)

    # Status validation
    dv2 = DataValidation(
        type="list", formula1='"Open,In Progress,Remediated,Closed"', allow_blank=True
    )
    dv2.sqref = f"K5:K{row - 1}"
    ws.add_data_validation(dv2)

    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = (
        "Severity Key: Critical = immediate legal/regulatory risk (30 days) | "
        "High = significant gap (90 days) | Medium = process improvement needed (180 days) | "
        "Low = enhancement opportunity (next audit cycle)"
    )
    ws[f"A{row}"].font = SMALL_FONT

    freeze_panes(ws, "A5")


def create_aggregation_sheet(wb):
    """Sheet 20: Aggregation & Conclusion."""
    ws = wb.create_sheet("20. Aggregation")
    set_col_widths(ws, [20, 30, 10, 10, 10, 10, 30, 35])

    ws.merge_cells("A1:H1")
    ws["A1"] = "Compliance Profile & Conclusion"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        "Aggregation of domain compliance assessments, principles assessment, "
        "A1727 rights compliance, and finding severity distribution."
    )
    ws["A2"].font = SMALL_FONT

    row = 4
    # --- Domain compliance profile ---
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Domain Compliance Profile"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1

    headers = [
        "Domain", "Name", "# C", "# PC", "# NC", "# NA",
        "Key Gaps", "Domain Narrative"
    ]
    write_row(ws, row, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1

    domain_data_start = row  # Track for totals formula
    for i, slug in enumerate(DOMAIN_ORDER):
        info = DOMAIN_INFO.get(slug, {})
        sheet_num = 3 + i
        domain_name = info.get("name", slug.replace("-", " ").title())
        sheet_title = f"{sheet_num}. {domain_name}"
        if len(sheet_title) > 31:
            sheet_title = sheet_title[:31]
        # Use single-quoted sheet reference for COUNTIF
        safe_title = sheet_title.replace("'", "''")
        # COUNTIF formulas referencing compliance column E in each domain sheet
        write_row(ws, row, [
            slug,
            info.get("name", slug),
            None, None, None, None, "", ""
        ])
        ws.cell(row=row, column=3).value = f"=COUNTIF('{safe_title}'!E:E,\"C\")"
        ws.cell(row=row, column=4).value = f"=COUNTIF('{safe_title}'!E:E,\"PC\")"
        ws.cell(row=row, column=5).value = f"=COUNTIF('{safe_title}'!E:E,\"NC\")"
        ws.cell(row=row, column=6).value = f"=COUNTIF('{safe_title}'!E:E,\"NA\")"
        row += 1

    # Totals row with SUM formulas
    domain_data_end = row - 1
    write_row(ws, row, ["", "TOTAL", None, None, None, None, "", ""])
    ws.cell(row=row, column=1).font = BOLD_BODY
    ws.cell(row=row, column=2).font = BOLD_BODY
    for col in range(3, 7):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col).value = f"=SUM({col_letter}{domain_data_start}:{col_letter}{domain_data_end})"
        ws.cell(row=row, column=col).font = BOLD_BODY
    row += 1

    # Compliance percentage row
    write_row(ws, row, ["", "% Compliant", None, None, None, None, "", ""])
    ws.cell(row=row, column=2).font = BOLD_BODY
    total_row = row - 1
    ws.cell(row=row, column=3).value = f'=IF(SUM(C{total_row}:F{total_row})>0,C{total_row}/SUM(C{total_row}:F{total_row}),0)'
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=3).font = BOLD_BODY
    row += 1

    # --- Principle compliance summary ---
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Principle Compliance Summary"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1

    principle_headers = ["Principle", "Status", "", "", "", "", "Notes", ""]
    write_row(ws, row, principle_headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    for principle_name, _slug, _desc, _domains in PDPA_PRINCIPLES:
        write_row(ws, row, [principle_name, "", "", "", "", "", "", ""])
        row += 1

    # --- A1727 rights summary ---
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "A1727 Enhanced Rights Compliance Summary"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1

    rights_headers = ["Right", "Status", "", "", "", "", "Notes", ""]
    write_row(ws, row, rights_headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1
    for right_name, section, _domain, _desc, _tests in A1727_RIGHTS:
        write_row(ws, row, [f"{right_name} ({section})", "", "", "", "", "", "", ""])
        row += 1

    # --- Finding summary ---
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Finding Summary by Severity"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    # Finding summary with COUNTIF formulas referencing Finding Register (sheet 19)
    finding_sheet = "19. Finding Register"
    safe_finding = finding_sheet.replace("'", "''")
    finding_summary = [
        ("Critical findings",  f"=COUNTIF('{safe_finding}'!D:D,\"Critical\")"),
        ("High findings",      f"=COUNTIF('{safe_finding}'!D:D,\"High\")"),
        ("Medium findings",    f"=COUNTIF('{safe_finding}'!D:D,\"Medium\")"),
        ("Low findings",       f"=COUNTIF('{safe_finding}'!D:D,\"Low\")"),
        ("Informational",      f"=COUNTIF('{safe_finding}'!D:D,\"Informational\")"),
        ("Total findings",     None),  # Will be a SUM formula
    ]
    finding_start = row
    for label, formula in finding_summary:
        write_row(ws, row, [label, "", "", "", "", "", "", ""])
        ws.cell(row=row, column=1).font = BOLD_BODY
        if formula:
            ws.cell(row=row, column=2).value = formula
        row += 1
    # Total = SUM of individual severity counts
    ws.cell(row=row - 1, column=2).value = f"=SUM(B{finding_start}:B{row - 2})"
    ws.cell(row=row - 1, column=2).font = BOLD_BODY

    # --- Overall opinion ---
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Overall Opinion"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    opinions = [
        ("Satisfactory (Memuaskan)",
         "No Critical/High findings. NC count zero/near-zero. Active compliance monitoring. A1727 rights implemented."),
        ("Needs Improvement (Perlu Penambahbaikan)",
         "One or more High findings, OR multiple Medium findings with systemic gaps. A1727 readiness incomplete."),
        ("Unsatisfactory (Tidak Memuaskan)",
         "One or more Critical findings, OR multiple High findings across core domains. "
         "Material non-compliance with PDPA principles. A1727 obligations not addressed."),
    ]
    for opinion, criteria in opinions:
        write_row(ws, row, [opinion, criteria, "", "", "", "", "", ""])
        row += 1

    row += 1
    write_row(ws, row, ["Selected opinion:", "", "", "", "", "", "", ""])
    ws.cell(row=row, column=1).font = BOLD_BODY
    row += 2

    # Narrative inputs
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Narrative Conclusion Inputs"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    row += 1
    narrative_items = [
        "Compliance posture summary (counts, %, strongest/weakest domains)",
        "Principle compliance assessment (7 principles overview)",
        "A1727 readiness assessment (new rights, breach notification, penalties)",
        "Risk assessment summary (finding distribution by severity)",
        "Key areas requiring attention and immediate action",
        "Prior audit follow-up (if applicable)",
        "Director/officer liability exposure assessment",
    ]
    for item in narrative_items:
        write_row(ws, row, [item, "", "", "", "", "", "", ""])
        ws.merge_cells(f"B{row}:H{row}")
        row += 1

    freeze_panes(ws, "A4")


def create_pbc_sheet(wb):
    """Sheet 21: Document Request (PBC)."""
    ws = wb.create_sheet("21. Document Request")
    set_col_widths(ws, [10, 22, 45, 16, 14, 18, 30])

    ws.merge_cells("A1:G1")
    ws["A1"] = "Pre-Assessment Document Request List (PBC)"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws["A2"] = "Essential items: request 2 weeks before fieldwork. Important items: obtained during fieldwork."
    ws["A2"].font = SMALL_FONT

    row = 4
    headers = [
        "Item #", "Domain", "Document / Evidence", "Priority\n(Essential/Important)",
        "Format", "Status\n(Provided/In Progress/N-A)", "Notes"
    ]
    write_row(ws, row, headers, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER)
    row += 1

    item_num = 1
    current_domain = ""
    for domain_slug, doc, priority in PBC_ITEMS:
        domain_name = DOMAIN_INFO.get(domain_slug, {}).get("name", domain_slug)
        if domain_slug != current_domain:
            current_domain = domain_slug
        write_row(ws, row, [
            f"PBC-{item_num:03d}", domain_name, doc, priority, "PDF/DOCX", "", ""
        ])
        item_num += 1
        row += 1

    # Priority validation
    dv = DataValidation(
        type="list", formula1='"Essential,Important"', allow_blank=True
    )
    dv.sqref = f"D5:D{row - 1}"
    ws.add_data_validation(dv)

    # Status validation
    dv2 = DataValidation(
        type="list", formula1='"Provided,In Progress,N-A"', allow_blank=True
    )
    dv2.sqref = f"F5:F{row - 1}"
    ws.add_data_validation(dv2)

    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = f"Total document requests: {item_num - 1}"
    ws[f"A{row}"].font = BOLD_BODY

    freeze_panes(ws, "A5")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()

    print("Building PDPA Data Protection Audit Work Program...")
    print(f"  Library: {len(CONTROLS_LIST)} controls across {len(DOMAIN_ORDER)} domains")
    print()

    create_cover_sheet(wb)
    print("  [1/21] Cover & Scope")

    create_ropa_register(wb)
    print("  [2/21] ROPA Register")

    # Domain worksheets (sheets 3-12)
    for i, domain_slug in enumerate(DOMAIN_ORDER):
        sheet_num = 3 + i
        domain_info = DOMAIN_INFO.get(domain_slug, {"name": domain_slug})
        controls = CONTROLS_BY_DOMAIN.get(domain_slug, [])
        create_domain_worksheet(wb, domain_slug, domain_info, controls, sheet_num)
    print(f"  [3-12] Domain Worksheets ({len(DOMAIN_ORDER)} domains)")

    create_principles_sheet(wb)
    print("  [13/21] 7 Principles Assessment")

    create_a1727_rights_sheet(wb)
    print("  [14/21] A1727 Rights Compliance")

    create_breach_notification_sheet(wb)
    print("  [15/21] Breach Notification Readiness")

    create_cross_border_sheet(wb)
    print("  [16/21] Cross-Border Transfer Assessment")

    create_director_liability_sheet(wb)
    print("  [17/21] Director Liability (s133A)")

    create_penalty_exposure_sheet(wb)
    print("  [18/21] Penalty Exposure Matrix")

    create_finding_register(wb)
    print("  [19/21] Finding Register")

    create_aggregation_sheet(wb)
    print("  [20/21] Aggregation & Conclusion")

    create_pbc_sheet(wb)
    print("  [21/21] Document Request (PBC)")

    # Save
    output_path = os.path.join(SCRIPT_DIR, "PDPA-DPA-Core-WorkProgram.xlsx")
    wb.save(output_path)
    print(f"\nSaved: {output_path}")

    # Copy to Tech-Audit if available
    tech_audit_path = os.path.expanduser("~/claude/tech-audit/PDPA")
    if os.path.isdir(tech_audit_path):
        ta_output = os.path.join(tech_audit_path, "PDPA-DPA-Core-WorkProgram.xlsx")
        wb.save(ta_output)
        print(f"Saved: {ta_output}")
    else:
        print(f"Note: Tech-Audit/PDPA not found at {tech_audit_path} — skipped copy")

    # Summary stats
    total_reqs = sum(len(reqs) for reqs in AWP_REQUIREMENTS.values())
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames:
        print(f"  - {s}")

    print(f"\nSummary:")
    print(f"  Controls: {len(CONTROLS_LIST)}")
    print(f"  Domains: {len(DOMAIN_ORDER)}")
    print(f"  AWP Requirements: {total_reqs}")
    print(f"  PBC Items: {len(PBC_ITEMS)}")
    print(f"  Penalty Matrix Entries: {len(PENALTY_MATRIX)}")
    print(f"  PDPA Principles: {len(PDPA_PRINCIPLES)}")
    print(f"  A1727 Rights: {len(A1727_RIGHTS)}")


if __name__ == "__main__":
    main()
